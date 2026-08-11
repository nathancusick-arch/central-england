"""Central England monthly test-purchase report generator.

Run with:
    streamlit run "Central England Report Generator.py"

The app accepts the new audit export, the previous LIVE workbook, and the
current store database.  It returns a formula-driven LIVE workbook and a
values-only client workbook for the month detected in the audit export.
"""

from __future__ import annotations

import io
import hashlib
import re
import zipfile
from collections import Counter, OrderedDict
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter


GENERATOR_VERSION = "2026.08.12.2"


REPORT_COLUMNS = [
    "Order", "Client", "Visit", "Site", "Order Deadline", "Responsibility",
    "Premises Name", "Address1", "Address2", "Address3", "City", "Post Code",
    "Submitted Date", "Approved Date", "Item to order", "Actual Visit Date",
    "Actual Visit Time", "AM / PM", "Pass-Fail", "Pass-Fail2", "Abort Reason",
    "Extra Site 1", "Extra Site 2", "Extra Site 3", "Extra Site 4", "Extra Site 5",
    "VISITORSEX", "What type of alcohol did you purchase?",
    "Please give details of the alcohol purchased (brand and size):",
    "Did you make the purchase on its own or as part of a larger shop?",
    "Did the operator ask your age?",
    "Did the operator ask for your ID during the transaction?",
    "Did the operator make eye contact with you during the transaction?",
    "If eye contact was made, when was it FIRST made?",
    "In your opinion, did the operator make an assessment of your age?",
    "Was the operator wearing a name badge?", "If they were, please state their name:",
    "Please accurately describe the operator that served you (include hair colour and style, build, height and any distinguishing features):",
    'Was there any "Challenge 25" signage visible in the till area?',
    'Was the operator wearing a "Challenge 25" Badge?', "OTHER VISIT DETAILS",
    "How many staff members were serving?",
    "Please comment on the overall service you received (include queue length and unattended tills):",
    "From the receipt, please enter the store name:",
    "Please enter the receipt number (#000000):", "Please enter the C number (C:000000):",
    "Please enter the T number (T:00):",
    "Please describe the location and positions of the store (i.e. names of the stores on either side):",
    "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:",
    "Please confirm below whether or not you were asked for ID:",
]

COLUMN_MAP = OrderedDict([
    ("Order", "order_internal_id"), ("Client", "client_name"),
    ("Visit", "internal_id"), ("Site", "site_internal_id"),
    ("Order Deadline", "end_date"), ("Responsibility", "responsibility"),
    ("Premises Name", "site_name"), ("Address1", "site_address_1"),
    ("Address2", "site_address_2"), ("Address3", "site_address_3"),
    ("City", None), ("Post Code", "site_post_code"),
    ("Submitted Date", "submitted_date"), ("Approved Date", "approval_date"),
    ("Item to order", "item_to_order"), ("Actual Visit Date", "date_of_visit"),
    ("Actual Visit Time", "time_of_visit"), ("AM / PM", None),
    ("Pass-Fail", "primary_result"), ("Pass-Fail2", "secondary_result"),
    ("Abort Reason", "Please detail why you were unable to conduct this audit:"),
    ("Extra Site 1", "site_code"), ("Extra Site 2", None), ("Extra Site 3", None),
    ("Extra Site 4", None), ("Extra Site 5", None), ("VISITORSEX", None),
    ("What type of alcohol did you purchase?", [
        "What type of E-cigarette product did you purchase/attempt to purchase?",
        "What type of alcohol did you try to purchase?",
    ]),
    ("Please give details of the alcohol purchased (brand and size):", [
        "Please give details of the e-cig product that you purchased:",
        "Please give details of the cigarettes that you purchased:",
        "Please give details of the alcohol that you purchased:",
    ]),
    ("Did you make the purchase on its own or as part of a larger shop?",
     "Did you make the purchase on its own or as part of a larger shop?"),
    ("Did the operator ask your age?", None),
    ("Did the operator ask for your ID during the transaction?",
     "Did the staff member who served you ask for ID?"),
    ("Did the operator make eye contact with you during the transaction?",
     "Did the staff member who served you make eye contact with you during the transaction?"),
    ("If eye contact was made, when was it FIRST made?", "When was eye contact first made?"),
    ("In your opinion, did the operator make an assessment of your age?",
     "Did the staff member who served you look at you long enough to assess your age?  "),
    ("Was the operator wearing a name badge?",
     "Was the staff member who served you wearing a name badge?"),
    ("If they were, please state their name:",
     "What was the name of the staff member who served you?"),
    ("Please accurately describe the operator that served you (include hair colour and style, build, height and any distinguishing features):",
     "Please accurately describe the staff member who served you:"),
    ('Was there any "Challenge 25" signage visible in the till area?',
     "Was there any generic 'Challenge 25' material visible from the till?"),
    ('Was the operator wearing a "Challenge 25" Badge?',
     "Was the staff member wearing a 'Challenge 25' badge?"),
    ("OTHER VISIT DETAILS", None),
    ("How many staff members were serving?", [
        "How many staff members were working on the tills?",
        "How staff members were working on the tills?",
    ]),
    ("Please comment on the overall service you received (include queue length and unattended tills):",
     "Please comment on the overall service you received:"),
    ("From the receipt, please enter the store name:",
     "From the top of the receipt, please enter the store name:"),
    ("Please enter the receipt number (#000000):",
     "Please enter the receipt number (#000000) from the receipt:"),
    ("Please enter the C number (C:000000):",
     "Please enter the C number (C:000000) from the receipt:"),
    ("Please enter the T number (T:00):",
     "Please enter the T number (T:00) from the receipt:"),
    ("Please describe the location and positions of the store (i.e. names of the stores on either side):", None),
    ("Please use this space to explain anything unusual about your visit or to clarify any detail of your report:",
     "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:"),
    ("Please confirm below whether or not you were asked for ID:", [
        "Please confirm below whether or not you were asked for ID:",
        "Please confirm whether or not you were asked for ID, and if so, at what point during the transaction ID was requested:",
    ]),
])

PUBLIC_SHEETS = [
    "Summary Data", "Store Performance", "Org Level Performance",
    "DOW-TOD Performance", "Performance over Time", "Performance over Time Chart",
]
REQUIRED_LIVE_SHEETS = [
    "Checks", "Input", "Region", "This Period", "R12M", *PUBLIC_SHEETS,
]
DATE_COLUMNS = {"Order Deadline", "Submitted Date", "Approved Date", "Actual Visit Date"}
TIME_COLUMNS = {"Actual Visit Time"}


class ReportGenerationError(ValueError):
    """A friendly validation error that can be shown in Streamlit."""


@dataclass
class HierarchyRow:
    code: str
    name: str
    area: str
    region: str
    operations_manager: str = ""
    site_internal_id: str = ""
    status: str = "Active"


@dataclass
class GenerationResult:
    live_bytes: bytes
    client_bytes: bytes
    zip_bytes: bytes
    live_name: str
    client_name: str
    report_month: date
    stats: dict[str, Any]
    warnings: list[str]


def _text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", _text(value)).strip().casefold()


def _normalise_code(value: Any) -> str:
    text = _text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def _excel_code(value: str) -> Any:
    return int(value) if value.isdigit() else value


def _parse_date(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().replace(tzinfo=None)


def _parse_time(value: Any) -> time | None:
    if value in (None, ""):
        return None
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)
    parsed = pd.to_datetime(_text(value), errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().time().replace(tzinfo=None)


def _month_start(value: date | datetime) -> date:
    return date(value.year, value.month, 1)


def _add_months(value: date, months: int) -> date:
    month_index = value.year * 12 + value.month - 1 + months
    return date(month_index // 12, month_index % 12 + 1, 1)


def _map_value(row: pd.Series, mapping: str | list[str] | None,
               column_lookup: dict[str, str]) -> str:
    if mapping is None:
        return ""
    source_columns = mapping if isinstance(mapping, list) else [mapping]
    values = []
    for column in source_columns:
        actual_column = column if column in row.index else column_lookup.get(_norm_header(column))
        if actual_column is not None:
            value = _text(row[actual_column])
            if value:
                values.append(value)
    return " | ".join(values)


def _typed_report_value(header: str, value: Any) -> Any:
    if header in DATE_COLUMNS:
        return _parse_date(value)
    if header in TIME_COLUMNS:
        return _parse_time(value)
    if header == "Extra Site 1":
        code = _normalise_code(value)
        return _excel_code(code) if code else None
    if header == "How many staff members were serving?":
        text = _text(value)
        if re.fullmatch(r"\d+", text):
            return int(text)
    return _text(value) if value is not None else ""


def map_audit_export(csv_bytes: bytes) -> tuple[list[dict[str, Any]], dict[str, int], list[str]]:
    try:
        frame = pd.read_csv(io.BytesIO(csv_bytes), dtype=str, encoding="utf-8-sig")
    except Exception as exc:
        raise ReportGenerationError(f"The audit export could not be read as CSV: {exc}") from exc

    required = {"item_to_order", "primary_result", "date_of_visit", "internal_id", "site_internal_id"}
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ReportGenerationError("The audit export is missing required columns: " + ", ".join(missing))

    item = frame["item_to_order"].fillna("").str.strip().str.casefold()
    result = frame["primary_result"].fillna("").str.strip().str.casefold()
    rapid_mask = item.eq("rapid delivery")
    abort_mask = result.eq("abort")
    excluded_mask = rapid_mask | abort_mask
    filtered = frame.loc[~excluded_mask].copy()
    column_lookup = {_norm_header(column): column for column in frame.columns}

    records: list[dict[str, Any]] = []
    for _, source_row in filtered.iterrows():
        record = {
            report_column: _typed_report_value(
                report_column, _map_value(source_row, source_mapping, column_lookup)
            )
            for report_column, source_mapping in COLUMN_MAP.items()
        }
        records.append(record)

    bad_dates = [r.get("Visit", "") for r in records if r.get("Actual Visit Date") is None]
    if bad_dates:
        raise ReportGenerationError(
            f"{len(bad_dates)} included audit row(s) have an invalid Actual Visit Date. "
            f"Example visit: {bad_dates[0] or '(blank)'}"
        )
    stats = {
        "export_rows": len(frame),
        "rapid_delivery_removed": int(rapid_mask.sum()),
        "aborts_removed": int(abort_mask.sum()),
        "included_rows": len(records),
    }
    optional_sources = {
        col for mapping in COLUMN_MAP.values() for col in (mapping if isinstance(mapping, list) else [mapping])
        if col is not None
    }
    available_normalised = {_norm_header(column) for column in frame.columns}
    absent_optional = sorted(column for column in optional_sources if _norm_header(column) not in available_normalised)
    warnings = []
    if absent_optional:
        warnings.append(
            f"{len(absent_optional)} optional mapped export column(s) were absent and were left blank."
        )
    return records, stats, warnings


def detect_report_month(records: list[dict[str, Any]]) -> tuple[date, list[str]]:
    if not records:
        raise ReportGenerationError("No rows remain after Rapid Delivery and abort visits are excluded.")
    months = Counter(_month_start(r["Actual Visit Date"]) for r in records)
    report_month, count = months.most_common(1)[0]
    warnings = []
    if len(months) > 1:
        other = ", ".join(f"{m:%B %Y} ({n})" for m, n in sorted(months.items()) if m != report_month)
        warnings.append(
            f"The export spans multiple visit months. {report_month:%B %Y} was selected from {count} rows; "
            f"other month(s): {other}. All included rows remain in This Period."
        )
    return report_month, warnings


def _load_live_workbook(workbook_bytes: bytes):
    try:
        # The legacy template contains stale cached links to source files that no
        # longer exist.  Excel attempts to repair those link-cache parts unless
        # they are deliberately omitted when the template is loaded.
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False, keep_links=False)
    except Exception as exc:
        raise ReportGenerationError(f"The previous LIVE workbook could not be read: {exc}") from exc
    missing = [name for name in REQUIRED_LIVE_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise ReportGenerationError(
            "The previous workbook is not a compatible LIVE report. Missing tab(s): " + ", ".join(missing)
        )
    return workbook


def read_store_database(store_bytes: bytes) -> tuple[OrderedDict[str, HierarchyRow], dict[str, str]]:
    try:
        workbook = load_workbook(io.BytesIO(store_bytes), data_only=True, read_only=True, keep_links=False)
    except Exception as exc:
        raise ReportGenerationError(f"The Store Database could not be read: {exc}") from exc
    if "Coop Database" not in workbook.sheetnames:
        raise ReportGenerationError("The Store Database does not contain a 'Coop Database' tab.")

    emails_lookup: dict[str, str] = {}
    if "Emails" in workbook.sheetnames:
        for row in workbook["Emails"].iter_rows(min_row=1, max_col=7, values_only=True):
            area_name, manager = _text(row[5]), _text(row[6])
            if area_name and manager:
                emails_lookup[area_name.casefold()] = manager

    sheet = workbook["Coop Database"]
    headers = {_norm_header(value): index for index, value in enumerate(next(sheet.iter_rows(max_row=1, max_col=24, values_only=True)))}
    needed = ["store code", "site internal id", "region", "area name", "store name"]
    missing = [name for name in needed if name not in headers]
    if missing:
        workbook.close()
        raise ReportGenerationError("The Store Database is missing column(s): " + ", ".join(missing))

    stores: OrderedDict[str, HierarchyRow] = OrderedDict()
    site_id_to_code: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, max_col=24, values_only=True):
        code = _normalise_code(row[headers["store code"]])
        if not code:
            continue
        if code in stores:
            workbook.close()
            raise ReportGenerationError(f"The Store Database contains duplicate store code {code}.")
        area_name = _text(row[headers["area name"]])
        operations_manager = _text(row[headers.get("area", -1)]) if "area" in headers else ""
        if not operations_manager and area_name:
            operations_manager = emails_lookup.get(area_name.casefold(), "")
        area = area_name or operations_manager or "Unmapped"
        region = _text(row[headers["region"]]) or "Unmapped"
        name = _text(row[headers["store name"]]) or f"Store {code}"
        site_id = _text(row[headers["site internal id"]])
        stores[code] = HierarchyRow(code, name, area, region, operations_manager, site_id)
        if site_id:
            site_id_to_code[site_id.casefold()] = code
    workbook.close()
    if not stores:
        raise ReportGenerationError("No stores were found on the Store Database's 'Coop Database' tab.")
    return stores, site_id_to_code


def _sheet_header_map(sheet, header_row: int, max_column: int = 50) -> dict[str, int]:
    result: dict[str, int] = {}
    for column in range(1, min(sheet.max_column, max_column) + 1):
        header = _norm_header(sheet.cell(header_row, column).value)
        if header and header not in result:
            result[header] = column
    return result


def _canonical_sheet_columns(sheet, header_row: int = 3) -> dict[str, int]:
    by_header = _sheet_header_map(sheet, header_row, 50)
    columns: dict[str, int] = {}
    for fallback, header in enumerate(REPORT_COLUMNS, start=1):
        normalised = _norm_header(header)
        columns[header] = by_header.get(normalised, fallback)
    return columns


def _last_data_row(sheet, key_column: int, first_row: int) -> int:
    last = first_row - 1
    for row in range(first_row, sheet.max_row + 1):
        if sheet.cell(row, key_column).value not in (None, ""):
            last = row
    return last


def extract_previous_hierarchy(workbook) -> OrderedDict[str, HierarchyRow]:
    sheet = workbook["Region"]
    stores: OrderedDict[str, HierarchyRow] = OrderedDict()
    for row in range(2, sheet.max_row + 1):
        code = _normalise_code(sheet.cell(row, 1).value)
        if not code:
            continue
        stores[code] = HierarchyRow(
            code=code,
            name=_text(sheet.cell(row, 4).value) or f"Store {code}",
            area=_text(sheet.cell(row, 5).value) or "Closed",
            region=_text(sheet.cell(row, 6).value) or "Closed",
            status="Previous",
        )
    return stores


def extract_rolling_records(workbook) -> list[dict[str, Any]]:
    sheet = workbook["R12M"]
    columns = _canonical_sheet_columns(sheet)
    last_row = _last_data_row(sheet, 1, 4)
    records: list[dict[str, Any]] = []
    for row in range(4, last_row + 1):
        record = {header: sheet.cell(row, column).value for header, column in columns.items()}
        for header in DATE_COLUMNS:
            record[header] = _parse_date(record.get(header))
        record["Actual Visit Time"] = _parse_time(record.get("Actual Visit Time"))
        record["Extra Site 1"] = _excel_code(_normalise_code(record.get("Extra Site 1"))) if _normalise_code(record.get("Extra Site 1")) else None
        records.append(record)
    return records


def _resolve_record_code(record: dict[str, Any], site_id_to_code: dict[str, str]) -> str:
    code = _normalise_code(record.get("Extra Site 1"))
    if not code:
        code = site_id_to_code.get(_text(record.get("Site")).casefold(), "")
        if code:
            record["Extra Site 1"] = _excel_code(code)
    return code


def build_rolling_records(
    previous_records: list[dict[str, Any]],
    current_records: list[dict[str, Any]],
    report_month: date,
) -> tuple[list[dict[str, Any]], int]:
    start_month = _add_months(report_month, -11)
    end_month = _add_months(report_month, 1)
    current_visits = {_text(record.get("Visit")) for record in current_records if _text(record.get("Visit"))}
    candidates = [
        record for record in previous_records
        if _text(record.get("Visit")) not in current_visits
    ] + current_records
    deduplicated: OrderedDict[str, dict[str, Any]] = OrderedDict()
    anonymous = 0
    for record in candidates:
        visit = _text(record.get("Visit"))
        key = visit or f"__anonymous_{anonymous}"
        anonymous += not bool(visit)
        deduplicated[key] = record

    rolling = []
    dropped = 0
    for record in deduplicated.values():
        visit_date = _parse_date(record.get("Actual Visit Date"))
        if visit_date and start_month <= visit_date.date() < end_month:
            record["Actual Visit Date"] = visit_date
            rolling.append(record)
        else:
            dropped += 1
    rolling.sort(key=lambda record: (
        record.get("Actual Visit Date") or datetime.min,
        _text(record.get("Visit")),
    ))
    return rolling, dropped


def build_hierarchy(
    database_stores: OrderedDict[str, HierarchyRow],
    previous_stores: OrderedDict[str, HierarchyRow],
    current_records: list[dict[str, Any]],
    rolling_records: list[dict[str, Any]],
    site_id_to_code: dict[str, str],
) -> tuple[OrderedDict[str, HierarchyRow], list[str]]:
    active_codes: OrderedDict[str, None] = OrderedDict()
    current_names: dict[str, str] = {}
    for record in [*rolling_records, *current_records]:
        code = _resolve_record_code(record, site_id_to_code)
        if code:
            active_codes.setdefault(code, None)
            if _text(record.get("Premises Name")):
                current_names[code] = _text(record.get("Premises Name")).removeprefix("Co-op ").strip(" ,")

    hierarchy: OrderedDict[str, HierarchyRow] = OrderedDict()
    for code in previous_stores:
        if code in database_stores:
            hierarchy[code] = database_stores[code]
        elif code in active_codes:
            old = previous_stores[code]
            hierarchy[code] = HierarchyRow(code, old.name, "Closed", "Closed", status="Closed")
    for code, store in database_stores.items():
        hierarchy.setdefault(code, store)
    warnings = []
    for code in active_codes:
        if code not in hierarchy:
            name = current_names.get(code, f"Store {code}")
            hierarchy[code] = HierarchyRow(code, name, "Unmapped", "Unmapped", status="Unmapped")
            warnings.append(f"Store {code} ({name}) appears in report data but not in the Store Database.")
    return hierarchy, warnings


def _capture_formula_template(sheet, row: int, start_column: int = 51) -> dict[int, str]:
    return {
        column: sheet.cell(row, column).value
        for column in range(start_column, sheet.max_column + 1)
        if isinstance(sheet.cell(row, column).value, str) and sheet.cell(row, column).value.startswith("=")
    }


def _copy_row_style(sheet, source_row: int, target_row: int, max_column: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _write_data_sheet(sheet, records: list[dict[str, Any]], filter_last_column: str) -> int:
    columns = _canonical_sheet_columns(sheet)
    old_last = _last_data_row(sheet, 1, 4)
    new_last = 3 + len(records)
    formula_template = _capture_formula_template(sheet, 4)
    max_column = max(sheet.max_column, max(formula_template, default=50))
    initial_max_row = sheet.max_row
    clear_to = max(old_last, new_last)
    for row in range(4, clear_to + 1):
        if row > initial_max_row:
            _copy_row_style(sheet, 4, row, max_column)
        for column in range(1, max_column + 1):
            sheet.cell(row, column).value = None
    for offset, record in enumerate(records):
        row = 4 + offset
        if row > sheet.max_row:
            _copy_row_style(sheet, 4, row, max_column)
        for header, column in columns.items():
            sheet.cell(row, column).value = record.get(header)
        for column, formula in formula_template.items():
            origin = f"{get_column_letter(column)}4"
            destination = f"{get_column_letter(column)}{row}"
            sheet.cell(row, column).value = Translator(formula, origin=origin).translate_formula(destination)
    count_end = max(4, new_last)
    sheet["A2"] = f"=COUNTA(A4:A{count_end})"
    sheet.auto_filter.ref = f"A3:{filter_last_column}{count_end}"
    return new_last


def _unique(values: Iterable[str], include: Iterable[str] = ()) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in [*values, *include]:
        clean = _text(value)
        if clean and clean.casefold() not in seen:
            seen.add(clean.casefold())
            result.append(clean)
    return result


def update_region_sheet(workbook, hierarchy: OrderedDict[str, HierarchyRow], org_labels: list[str]) -> int:
    sheet = workbook["Region"]
    old_last = _last_data_row(sheet, 1, 2)
    new_last = 1 + len(hierarchy)
    for row in range(2, max(old_last, new_last) + 1):
        for column in range(1, 8):
            sheet.cell(row, column).value = None
    for row, store in enumerate(hierarchy.values(), start=2):
        code = _excel_code(store.code)
        values = [code, code, code, store.name, store.area, store.region, store.name]
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column).value = value
    sheet.auto_filter.ref = f"A1:G{new_last}"

    # Refresh the small management/check lists alongside the store mapping.
    for row in range(1, max(sheet.max_row, len(org_labels) + 2) + 1):
        for column in [12, 13, 15, 18]:
            sheet.cell(row, column).value = None
    operations = _unique(store.operations_manager for store in hierarchy.values())
    regions = _unique(store.region for store in hierarchy.values() if store.region not in {"Closed", "Unmapped"})
    sheet["M1"] = "Org Level"
    sheet["O1"] = "Operations Manager"
    sheet["R1"] = "Group Manager"
    for row, label in enumerate(org_labels, start=2):
        sheet.cell(row, 13).value = label
        sheet.cell(row, 12).value = f"=IFERROR(VLOOKUP(M{row},'Org Level Performance'!A:A,1,FALSE),\"Missing\")"
    for row, manager in enumerate(operations, start=2):
        sheet.cell(row, 15).value = manager
    for row, manager in enumerate(regions, start=2):
        sheet.cell(row, 18).value = manager
    return new_last


def _countif_formula(data_sheet: str, key_column: str, data_end: int, label_cell: str, status: str) -> str:
    return f'=COUNTIF(\'{data_sheet}\'!${key_column}$4:${key_column}${data_end},{label_cell}&"{status}")'


def _write_metric_formulas(sheet, row: int, label_cell: str, current_col: str, rolling_col: str,
                           current_end: int, rolling_end: int, start_column: int = 2) -> None:
    current = {status: _countif_formula("This Period", current_col, current_end, label_cell, status)
               for status in ["PASS", "FAIL", "ABORT"]}
    rolling = {status: _countif_formula("R12M", rolling_col, rolling_end, label_cell, status)
               for status in ["PASS", "FAIL", "ABORT"]}
    current_completed = get_column_letter(start_column + 1)
    current_passes = get_column_letter(start_column + 3)
    rolling_completed = get_column_letter(start_column + 6)
    rolling_passes = get_column_letter(start_column + 8)
    formulas = [
        f"={current['PASS'][1:]}+{current['FAIL'][1:]}+{current['ABORT'][1:]}",
        f"={current['PASS'][1:]}+{current['FAIL'][1:]}", current["FAIL"], current["PASS"],
        f'=IF({current_completed}{row}=0,"-",{current_passes}{row}/{current_completed}{row})',
        f"={rolling['PASS'][1:]}+{rolling['FAIL'][1:]}+{rolling['ABORT'][1:]}",
        f"={rolling['PASS'][1:]}+{rolling['FAIL'][1:]}", rolling["FAIL"], rolling["PASS"],
        f'=IF({rolling_completed}{row}=0,"-",{rolling_passes}{row}/{rolling_completed}{row})',
    ]
    for column, formula in enumerate(formulas, start=start_column):
        sheet.cell(row, column).value = formula


def update_store_performance(workbook, hierarchy: OrderedDict[str, HierarchyRow],
                             current_end: int, rolling_end: int) -> int:
    sheet = workbook["Store Performance"]
    total_row = next((row for row in range(7, sheet.max_row + 1) if _text(sheet.cell(row, 1).value).casefold() == "total"), 334)
    capacity = total_row - 8
    if len(hierarchy) > capacity:
        extra = len(hierarchy) - capacity
        sheet.insert_rows(total_row - 1, amount=extra)
        for row in range(total_row - 1, total_row - 1 + extra):
            _copy_row_style(sheet, 7, row, 12)
        total_row += extra
    for row in range(7, total_row):
        for column in range(1, 13):
            sheet.cell(row, column).value = None
    for row, store in enumerate(hierarchy.values(), start=7):
        sheet.cell(row, 1).value = _excel_code(store.code)
        sheet.cell(row, 2).value = store.name
        _write_metric_formulas(
            sheet, row, f"'Store Performance'!$A{row}", "IO", "IS",
            current_end, rolling_end, start_column=3,
        )
    data_end = 6 + len(hierarchy)
    desired_total_row = data_end + 2
    if desired_total_row < total_row:
        _copy_row_style(sheet, total_row, desired_total_row, 12)
        sheet.delete_rows(desired_total_row + 1, amount=total_row - desired_total_row)
        total_row = desired_total_row
    sheet.cell(total_row, 1).value = "Total"
    for column in [3, 4, 5, 6, 8, 9, 10, 11]:
        letter = get_column_letter(column)
        sheet.cell(total_row, column).value = f"=SUM({letter}7:{letter}{data_end})"
    sheet.cell(total_row, 7).value = f'=IF(D{total_row}=0,"-",F{total_row}/D{total_row})'
    sheet.cell(total_row, 12).value = f'=IF(I{total_row}=0,"-",K{total_row}/I{total_row})'
    sheet["L2"] = "='Summary Data'!AA2"
    sheet["L3"] = "=Checks!B19"
    sheet.auto_filter.ref = f"A6:L{data_end}"
    workbook["Checks"]["G11"] = f"='Store Performance'!H{total_row}"
    workbook["Checks"]["H11"] = f"='Store Performance'!I{total_row}"
    workbook["Checks"]["I11"] = f"='Store Performance'!K{total_row}"
    workbook["Checks"]["J11"] = f"='Store Performance'!J{total_row}"
    workbook["Checks"]["K11"] = f"='Store Performance'!L{total_row}"
    return total_row


def _ensure_section_capacity(sheet, start_row: int, total_row: int, needed: int, style_row: int) -> int:
    capacity = total_row - start_row - 1
    if needed <= capacity:
        return total_row
    extra = needed - capacity
    sheet.insert_rows(total_row, amount=extra)
    for row in range(total_row, total_row + extra):
        _copy_row_style(sheet, style_row, row, 11)
    return total_row + extra


def _write_org_section(sheet, labels: list[str], start_row: int, total_row: int,
                       current_col: str, rolling_col: str, current_end: int, rolling_end: int) -> int:
    total_row = _ensure_section_capacity(sheet, start_row, total_row, len(labels), max(start_row, total_row - 2))
    for row in range(start_row, total_row):
        for column in range(1, 12):
            sheet.cell(row, column).value = None
    for row, label in enumerate(labels, start=start_row):
        sheet.cell(row, 1).value = label
        _write_metric_formulas(sheet, row, f"'Org Level Performance'!$A{row}", current_col, rolling_col,
                               current_end, rolling_end)
    data_end = start_row + len(labels) - 1
    sheet.cell(total_row, 1).value = "Total"
    for column in [2, 3, 4, 5, 7, 8, 9, 10]:
        letter = get_column_letter(column)
        sheet.cell(total_row, column).value = f"=SUM({letter}{start_row}:{letter}{data_end})"
    sheet.cell(total_row, 6).value = f'=IF(C{total_row}=0,"-",E{total_row}/C{total_row})'
    sheet.cell(total_row, 11).value = f'=IF(H{total_row}=0,"-",J{total_row}/H{total_row})'
    return total_row


def update_org_performance(workbook, area_labels: list[str], region_labels: list[str],
                           current_end: int, rolling_end: int) -> tuple[int, int]:
    sheet = workbook["Org Level Performance"]
    first_total = next(row for row in range(7, sheet.max_row + 1)
                       if _text(sheet.cell(row, 1).value).casefold() == "total")
    old_first_total = first_total
    first_total = _write_org_section(sheet, area_labels, 7, first_total, "IB", "IE", current_end, rolling_end)
    shift = first_total - old_first_total
    second_header = 28 + shift
    second_start = second_header + 1
    second_total = next(row for row in range(second_start, sheet.max_row + 1)
                        if _text(sheet.cell(row, 1).value).casefold() == "total")
    second_total = _write_org_section(sheet, region_labels, second_start, second_total, "IA", "IF",
                                      current_end, rolling_end)
    sheet["K2"] = "='Summary Data'!AA2"
    sheet["K3"] = "=Checks!B19"
    workbook["Checks"]["G12"] = f"='Org Level Performance'!G{first_total}"
    workbook["Checks"]["H12"] = f"='Org Level Performance'!H{first_total}"
    workbook["Checks"]["I12"] = f"='Org Level Performance'!J{first_total}"
    workbook["Checks"]["J12"] = f"='Org Level Performance'!I{first_total}"
    workbook["Checks"]["K12"] = f"='Org Level Performance'!K{first_total}"
    return first_total, second_total


def update_dow_tod_performance(workbook, current_end: int, rolling_end: int) -> None:
    sheet = workbook["DOW-TOD Performance"]
    for row in range(7, 14):
        _write_metric_formulas(sheet, row, f"'DOW-TOD Performance'!$A{row}", "IL", "IP", current_end, rolling_end)
    for row in range(20, 23):
        _write_metric_formulas(sheet, row, f"'DOW-TOD Performance'!$A{row}", "IE", "II", current_end, rolling_end)
    for total_row, start_row, end_row in [(15, 7, 13), (24, 20, 22)]:
        for column in [2, 3, 4, 5, 7, 8, 9, 10]:
            letter = get_column_letter(column)
            sheet.cell(total_row, column).value = f"=SUM({letter}{start_row}:{letter}{end_row})"
        sheet.cell(total_row, 6).value = f'=IF(C{total_row}=0,"-",E{total_row}/C{total_row})'
        sheet.cell(total_row, 11).value = f'=IF(H{total_row}=0,"-",J{total_row}/H{total_row})'
    sheet["K2"] = "='Summary Data'!AA2"
    sheet["K3"] = "=Checks!B19"


def update_summary_live(workbook, current_count: int) -> None:
    sheet = workbook["Summary Data"]
    old_last = _last_data_row(sheet, 1, 8)
    new_last = 7 + current_count
    formula_template = {
        column: sheet.cell(8, column).value
        for column in range(1, 28)
        if isinstance(sheet.cell(8, column).value, str) and sheet.cell(8, column).value.startswith("=")
    }
    initial_max_row = sheet.max_row
    for row in range(8, max(old_last, new_last) + 1):
        for column in range(1, 28):
            sheet.cell(row, column).value = None
    for row in range(8, new_last + 1):
        if row > initial_max_row:
            _copy_row_style(sheet, 8, row, 27)
        for column, formula in formula_template.items():
            origin = f"{get_column_letter(column)}8"
            destination = f"{get_column_letter(column)}{row}"
            sheet.cell(row, column).value = Translator(formula, origin=origin).translate_formula(destination)
    sheet["B3"] = f"=COUNTA($A$8:$A${new_last})"
    sheet["B4"] = f'=COUNTIF($G$8:$G${new_last},"pass")+COUNTIF($G$8:$G${new_last},"fail")'
    sheet["B5"] = f'=IFERROR(COUNTIF($G$8:$G${new_last},"pass")/$B$4,"-")'
    sheet["AA2"] = "Central England Co-Op"


def update_performance_over_time_live(workbook, report_month: date, rolling_end: int) -> None:
    sheet = workbook["Performance over Time"]
    for offset, column in enumerate(range(2, 14), start=-11):
        period = _add_months(report_month, offset)
        sheet.cell(5, column).value = datetime(period.year, period.month, 1)
        sheet.cell(5, column).number_format = "dd/mm/yyyy"
        letter = get_column_letter(column)
        pass_count = (
            f'COUNTIFS(\'R12M\'!$P$4:$P${rolling_end},">="&{letter}$5,'
            f'\'R12M\'!$P$4:$P${rolling_end},"<"&EDATE({letter}$5,1),'
            f'\'R12M\'!$S$4:$S${rolling_end},"pass")'
        )
        fail_count = (
            f'COUNTIFS(\'R12M\'!$P$4:$P${rolling_end},">="&{letter}$5,'
            f'\'R12M\'!$P$4:$P${rolling_end},"<"&EDATE({letter}$5,1),'
            f'\'R12M\'!$S$4:$S${rolling_end},"fail")'
        )
        sheet.cell(6, column).value = f'=IFERROR({pass_count}/({pass_count}+{fail_count}),"-")'
    sheet["K2"] = "='Summary Data'!AA2"


def update_report_month(workbook, report_month: date) -> None:
    checks = workbook["Checks"]
    checks["B19"] = datetime(report_month.year, report_month.month, 1)
    checks["B19"].number_format = "mmmm yy"
    # Repair the old fixed This Period order-count ranges.
    input_sheet = workbook["Input"]
    actual_end = max(4, _last_data_row(workbook["This Period"], 1, 4))
    for row in range(17, input_sheet.max_row + 1):
        if input_sheet.cell(row, 1).value:
            input_sheet.cell(row, 6).value = f'=COUNTIF(\'This Period\'!$A$4:$A${actual_end},Input!$A{row})'


def _record_result(record: dict[str, Any]) -> str:
    return _text(record.get("Pass-Fail")).casefold()


def _record_bucket(record: dict[str, Any], hierarchy: OrderedDict[str, HierarchyRow], attribute: str) -> str:
    code = _normalise_code(record.get("Extra Site 1"))
    store = hierarchy.get(code)
    return getattr(store, attribute) if store else "Unmapped"


def _time_bucket(value: Any) -> str:
    parsed = _parse_time(value)
    if parsed is None:
        return "Night"
    minutes = parsed.hour * 60 + parsed.minute
    if 7 * 60 <= minutes < 12 * 60:
        return "Morning"
    if 12 * 60 <= minutes < 17 * 60:
        return "Afternoon"
    return "Night"


def _day_bucket(value: Any) -> str:
    parsed = _parse_date(value)
    if parsed is None:
        return ""
    return ["Mon", "Tues", "Wed", "Thur", "Fri", "Sat", "Sun"][parsed.weekday()]


def _metrics(records: Iterable[dict[str, Any]], key_function, labels: Iterable[str]) -> dict[str, list[Any]]:
    counters = {label: Counter() for label in labels}
    for record in records:
        label = key_function(record)
        if label not in counters:
            counters[label] = Counter()
        result = _record_result(record)
        if result in {"pass", "fail", "abort"}:
            counters[label][result] += 1
    output = {}
    for label, counts in counters.items():
        completed = counts["pass"] + counts["fail"]
        visits = completed + counts["abort"]
        output[label] = [visits, completed, counts["fail"], counts["pass"], counts["pass"] / completed if completed else "-"]
    return output


def _summary_values(record: dict[str, Any]) -> list[Any]:
    visit_time = _parse_time(record.get("Actual Visit Time"))
    return [
        record.get("Extra Site 1") or "", _text(record.get("Premises Name")), _text(record.get("Post Code")),
        _parse_date(record.get("Actual Visit Date")), _time_bucket(visit_time),
        visit_time.strftime("%H:%M ") if visit_time else "", _record_result(record),
        _text(record.get("If eye contact was made, when was it FIRST made?")), "",
        _text(record.get("In your opinion, did the operator make an assessment of your age?")),
        _text(record.get("Did the operator ask for your ID during the transaction?")),
        _text(record.get("Was the operator wearing a name badge?")),
        _text(record.get('Was the operator wearing a "Challenge 25" Badge?')),
        _text(record.get("What type of alcohol did you purchase?")),
        _text(record.get("Please give details of the alcohol purchased (brand and size):")),
        _text(record.get("Did you make the purchase on its own or as part of a larger shop?")),
        _text(record.get("Did the operator make eye contact with you during the transaction?")),
        _text(record.get("If they were, please state their name:")),
        _text(record.get("Please accurately describe the operator that served you (include hair colour and style, build, height and any distinguishing features):")),
        _text(record.get('Was there any "Challenge 25" signage visible in the till area?')),
        record.get("How many staff members were serving?") or "",
        _text(record.get("Please comment on the overall service you received (include queue length and unattended tills):")),
        _text(record.get("From the receipt, please enter the store name:")),
        record.get("Please enter the receipt number (#000000):") or "",
        record.get("Please enter the C number (C:000000):") or "",
        record.get("Please enter the T number (T:00):") or "",
        _text(record.get("Please use this space to explain anything unusual about your visit or to clarify any detail of your report:")),
    ]


def populate_static_reports(workbook, current_records: list[dict[str, Any]], rolling_records: list[dict[str, Any]],
                            hierarchy: OrderedDict[str, HierarchyRow], area_labels: list[str],
                            region_labels: list[str], report_month: date,
                            store_total_row: int, org_total_rows: tuple[int, int]) -> None:
    period_label = report_month.strftime("%B %y")
    company = "Central England Co-Op"

    summary = workbook["Summary Data"]
    summary["B3"] = len(current_records)
    summary["B4"] = sum(_record_result(record) in {"pass", "fail"} for record in current_records)
    passes = sum(_record_result(record) == "pass" for record in current_records)
    summary["B5"] = passes / summary["B4"].value if summary["B4"].value else "-"
    summary["AA2"] = company
    old_last = _last_data_row(summary, 1, 8)
    new_last = 7 + len(current_records)
    for row in range(8, max(old_last, new_last) + 1):
        for column in range(1, 28):
            summary.cell(row, column).value = None
    for row, record in enumerate(current_records, start=8):
        for column, value in enumerate(_summary_values(record), start=1):
            summary.cell(row, column).value = value
    if summary.max_row > new_last:
        summary.delete_rows(new_last + 1, amount=summary.max_row - new_last)

    store_sheet = workbook["Store Performance"]
    store_sheet["L2"] = company
    store_sheet["L3"] = period_label
    current_store = _metrics(current_records, lambda r: _normalise_code(r.get("Extra Site 1")), hierarchy.keys())
    rolling_store = _metrics(rolling_records, lambda r: _normalise_code(r.get("Extra Site 1")), hierarchy.keys())
    for row, store in enumerate(hierarchy.values(), start=7):
        store_sheet.cell(row, 1).value = _excel_code(store.code)
        store_sheet.cell(row, 2).value = store.name
        values = [*current_store.get(store.code, [0, 0, 0, 0, "-"]), *rolling_store.get(store.code, [0, 0, 0, 0, "-"])]
        for column, value in enumerate(values, start=3):
            store_sheet.cell(row, column).value = value
    store_current_total = _metrics(current_records, lambda _: "Total", ["Total"])["Total"]
    store_rolling_total = _metrics(rolling_records, lambda _: "Total", ["Total"])["Total"]
    for column, value in enumerate([*store_current_total, *store_rolling_total], start=3):
        store_sheet.cell(store_total_row, column).value = value

    org_sheet = workbook["Org Level Performance"]
    org_sheet["K2"] = company
    org_sheet["K3"] = period_label
    first_total, second_total = org_total_rows
    for labels, start_row, total_row, attribute in [
        (area_labels, 7, first_total, "area"),
        (region_labels, first_total + 5, second_total, "region"),
    ]:
        current_values = _metrics(current_records, lambda r, a=attribute: _record_bucket(r, hierarchy, a), labels)
        rolling_values = _metrics(rolling_records, lambda r, a=attribute: _record_bucket(r, hierarchy, a), labels)
        for row, label in enumerate(labels, start=start_row):
            org_sheet.cell(row, 1).value = label
            values = [*current_values[label], *rolling_values[label]]
            for column, value in enumerate(values, start=2):
                org_sheet.cell(row, column).value = value
        current_total = _metrics(current_records, lambda _: "Total", ["Total"])["Total"]
        rolling_total = _metrics(rolling_records, lambda _: "Total", ["Total"])["Total"]
        for column, value in enumerate([*current_total, *rolling_total], start=2):
            org_sheet.cell(total_row, column).value = value

    dow_sheet = workbook["DOW-TOD Performance"]
    dow_sheet["K2"] = company
    dow_sheet["K3"] = period_label
    for labels, start_row, total_row, key_function in [
        (["Mon", "Tues", "Wed", "Thur", "Fri", "Sat", "Sun"], 7, 15,
         lambda r: _day_bucket(r.get("Actual Visit Date"))),
        (["Morning", "Afternoon", "Night"], 20, 24,
         lambda r: _time_bucket(r.get("Actual Visit Time"))),
    ]:
        current_values = _metrics(current_records, key_function, labels)
        rolling_values = _metrics(rolling_records, key_function, labels)
        for row, label in enumerate(labels, start=start_row):
            values = [*current_values[label], *rolling_values[label]]
            for column, value in enumerate(values, start=2):
                dow_sheet.cell(row, column).value = value
        current_total = _metrics(current_records, lambda _: "Total", ["Total"])["Total"]
        rolling_total = _metrics(rolling_records, lambda _: "Total", ["Total"])["Total"]
        for column, value in enumerate([*current_total, *rolling_total], start=2):
            dow_sheet.cell(total_row, column).value = value

    trend = workbook["Performance over Time"]
    trend["K2"] = company
    for offset, column in enumerate(range(2, 14), start=-11):
        month = _add_months(report_month, offset)
        trend.cell(5, column).value = datetime(month.year, month.month, 1)
        trend.cell(5, column).number_format = "dd/mm/yyyy"
        next_month = _add_months(month, 1)
        month_records = [
            record for record in rolling_records
            if (visit_date := _parse_date(record.get("Actual Visit Date")))
            and month <= visit_date.date() < next_month
        ]
        trend.cell(6, column).value = _metrics(month_records, lambda _: "Month", ["Month"])["Month"][4]

    # The client workbook must contain no live formulas or internal tabs.
    for sheet_name in PUBLIC_SHEETS:
        sheet = workbook[sheet_name]
        if not hasattr(sheet, "iter_rows"):  # The chart-only tab is an Excel Chartsheet.
            continue
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.value = None


def _save_workbook(workbook) -> bytes:
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _remove_invalid_defined_names(workbook, removed_sheets: set[str] | None = None) -> None:
    """Remove legacy names that would make Excel repair workbook.xml."""
    removed_sheets = removed_sheets or set()

    def is_invalid(defined_name) -> bool:
        reference = _text(getattr(defined_name, "attr_text", ""))
        if "#REF!" in reference or re.search(r"\[\d+\]", reference):
            return True
        return any(
            f"'{sheet_name}'!" in reference or f"{sheet_name}!" in reference
            for sheet_name in removed_sheets
        )

    for name, defined_name in list(workbook.defined_names.items()):
        if is_invalid(defined_name):
            del workbook.defined_names[name]
    for sheet in workbook.worksheets:
        for name, defined_name in list(sheet.defined_names.items()):
            if is_invalid(defined_name):
                del sheet.defined_names[name]


def _remove_external_formulas(workbook) -> None:
    """Discard unused legacy formulas whose source workbooks are not included."""
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" and re.search(r"\[\d+\]", str(cell.value)):
                    cell.value = None


def _select_opening_sheet(workbook, title: str) -> None:
    """Make one worksheet the unambiguous tab selected when Excel opens."""
    for sheet in workbook.worksheets:
        sheet.sheet_view.tabSelected = False
    target = workbook[title]
    workbook.active = workbook._sheets.index(target)
    target.sheet_view.tabSelected = True


def generate_reports(csv_bytes: bytes, previous_live_bytes: bytes, store_bytes: bytes) -> GenerationResult:
    current_records, stats, warnings = map_audit_export(csv_bytes)
    report_month, month_warnings = detect_report_month(current_records)
    warnings.extend(month_warnings)
    workbook = _load_live_workbook(previous_live_bytes)
    previous_hierarchy = extract_previous_hierarchy(workbook)
    previous_rolling = extract_rolling_records(workbook)
    database_stores, site_id_to_code = read_store_database(store_bytes)
    rolling_records, dropped = build_rolling_records(previous_rolling, current_records, report_month)
    hierarchy, hierarchy_warnings = build_hierarchy(
        database_stores, previous_hierarchy, current_records, rolling_records, site_id_to_code
    )
    warnings.extend(hierarchy_warnings)

    statuses = {store.status for store in hierarchy.values()}
    area_labels = _unique(
        (store.area for store in hierarchy.values() if store.status == "Active"),
        include=[label for label in ["Closed", "Unmapped"] if label in statuses],
    )
    region_labels = _unique(
        (store.region for store in hierarchy.values() if store.status == "Active"),
        include=[label for label in ["Closed", "Unmapped"] if label in statuses],
    )

    current_end = _write_data_sheet(workbook["This Period"], current_records, "IR")
    rolling_end = _write_data_sheet(workbook["R12M"], rolling_records, "JS")
    update_region_sheet(workbook, hierarchy, _unique([*area_labels, *region_labels]))
    update_summary_live(workbook, len(current_records))
    store_total_row = update_store_performance(workbook, hierarchy, current_end, rolling_end)
    org_total_rows = update_org_performance(workbook, area_labels, region_labels, current_end, rolling_end)
    update_dow_tod_performance(workbook, current_end, rolling_end)
    update_performance_over_time_live(workbook, report_month, rolling_end)
    update_report_month(workbook, report_month)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook._external_links = []
    _remove_external_formulas(workbook)
    _remove_invalid_defined_names(workbook)
    live_bytes = _save_workbook(workbook)
    workbook.close()

    client_workbook = load_workbook(io.BytesIO(live_bytes), data_only=False, keep_links=False)
    removed_sheets = {
        sheet_name for sheet_name in client_workbook.sheetnames
        if sheet_name not in PUBLIC_SHEETS
    }
    for sheet_name in list(client_workbook.sheetnames):
        if sheet_name not in PUBLIC_SHEETS:
            client_workbook.remove(client_workbook[sheet_name])
    populate_static_reports(
        client_workbook, current_records, rolling_records, hierarchy, area_labels, region_labels,
        report_month, store_total_row, org_total_rows,
    )
    client_workbook._external_links = []
    _remove_invalid_defined_names(client_workbook, removed_sheets)
    _select_opening_sheet(client_workbook, "Summary Data")
    client_bytes = _save_workbook(client_workbook)
    client_workbook.close()

    report_name = f"Central England Test Purchases Report - {report_month:%B %Y}"
    live_name = f"{report_name} LIVE.xlsx"
    client_name = f"{report_name}.xlsx"
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(live_name, live_bytes)
        archive.writestr(client_name, client_bytes)

    current_completed = [record for record in current_records if _record_result(record) in {"pass", "fail"}]
    rolling_completed = [record for record in rolling_records if _record_result(record) in {"pass", "fail"}]
    current_passes = sum(_record_result(record) == "pass" for record in current_completed)
    rolling_passes = sum(_record_result(record) == "pass" for record in rolling_completed)
    stats.update({
        "rolling_rows": len(rolling_records),
        "rolling_rows_dropped": dropped,
        "database_stores": len(database_stores),
        "report_stores": len(hierarchy),
        "closed_history_stores": sum(store.status == "Closed" for store in hierarchy.values()),
        "unmapped_stores": sum(store.status == "Unmapped" for store in hierarchy.values()),
        "completed_visits": len(current_completed),
        "pass_rate": current_passes / len(current_completed) if current_completed else None,
        "rolling_completed_visits": len(rolling_completed),
        "rolling_pass_rate": rolling_passes / len(rolling_completed) if rolling_completed else None,
    })
    return GenerationResult(
        live_bytes, client_bytes, zip_buffer.getvalue(), live_name, client_name,
        report_month, stats, warnings,
    )


def _build_email_text(generated: GenerationResult) -> str:
    pass_rate = generated.stats.get("pass_rate")
    rolling_pass_rate = generated.stats.get("rolling_pass_rate")
    pass_rate_text = f"{round(pass_rate * 100)}%" if pass_rate is not None else "N/A"
    rolling_pass_rate_text = (
        f"{round(rolling_pass_rate * 100)}%" if rolling_pass_rate is not None else "N/A"
    )
    return (
        "Hi All,\n\n"
        "Please find attached the Serve Legal report detailing the visits completed in "
        f"{generated.report_month:%B}.\n\n"
        f"As you’ll see from the report, the pass rate was {pass_rate_text} based on "
        f"{generated.stats['completed_visits']} completed visits. Your Rolling 12-month pass rate "
        f"currently stands at {rolling_pass_rate_text}."
    )


def run_app() -> None:
    import streamlit as st

    st.set_page_config(page_title="Central England Report Generator", page_icon="📊", layout="wide")
    st.title("Central England Report Generator")
    st.caption(f"Generator version {GENERATOR_VERSION}")
    st.write(
        "Upload the new monthly audit export, the previous month's LIVE report, and the current "
        "Store Database. The generator rolls the report forward and produces both delivery files."
    )
    with st.expander("What the generator updates"):
        st.markdown(
            "- Filters Rapid Delivery visits and aborts, matching the former mapper.\n"
            "- Replaces **This Period** with the new export.\n"
            "- Adds the new month to **R12M**, removes the expired month, and de-duplicates visits.\n"
            "- Refreshes store names, areas, regional management, new sites, and active historical closures from the Store Database.\n"
            "- Produces a formula-driven **LIVE** workbook and a six-tab, values-only client workbook."
        )

    col1, col2, col3 = st.columns(3)
    with col1:
        audit_upload = st.file_uploader("1. New audit data export", type=["csv"])
    with col2:
        live_upload = st.file_uploader("2. Previous LIVE report", type=["xlsx"])
    with col3:
        store_upload = st.file_uploader("3. Current Store Database", type=["xlsm", "xlsx"])

    ready = audit_upload is not None and live_upload is not None and store_upload is not None
    upload_payloads = None
    upload_signature = None
    if ready:
        upload_payloads = (
            audit_upload.getvalue(), live_upload.getvalue(), store_upload.getvalue()
        )
        upload_signature = tuple(
            (upload.name, len(payload), hashlib.sha256(payload).hexdigest())
            for upload, payload in zip(
                (audit_upload, live_upload, store_upload), upload_payloads
            )
        )

    if st.session_state.get("ce_report_input_signature") != upload_signature:
        st.session_state.pop("ce_generated_report", None)
        st.session_state["ce_report_input_signature"] = upload_signature

    if st.button("Generate reports", type="primary", disabled=not ready, use_container_width=True):
        try:
            with st.spinner("Generating and validating the new reports..."):
                generated = generate_reports(*upload_payloads)
            st.session_state["ce_generated_report"] = generated
        except ReportGenerationError as exc:
            st.session_state.pop("ce_generated_report", None)
            st.error(str(exc))
        except Exception as exc:
            st.session_state.pop("ce_generated_report", None)
            st.exception(exc)

    generated = st.session_state.get("ce_generated_report")
    if generated is not None:
        st.success(f"{generated.report_month:%B %Y} reports generated successfully.")
        metrics = st.columns(5)
        labels = [
            ("Included visits", "included_rows"), ("Rolling visits", "rolling_rows"),
            ("Store DB sites", "database_stores"), ("Report sites", "report_stores"),
            ("Unmapped sites", "unmapped_stores"),
        ]
        for column, (label, key) in zip(metrics, labels):
            column.metric(label, generated.stats[key])
        st.caption(
            f"Removed {generated.stats['rapid_delivery_removed']} Rapid Delivery row(s) and "
            f"{generated.stats['aborts_removed']} abort row(s). "
            f"Dropped {generated.stats['rolling_rows_dropped']} row(s) outside the new rolling window."
        )
        for warning in generated.warnings:
            st.warning(warning)
        st.download_button(
            "Download both reports (.zip)", generated.zip_bytes,
            file_name=f"Central England Reports - {generated.report_month:%B %Y}.zip",
            mime="application/zip", type="primary", use_container_width=True,
        )
        left, right = st.columns(2)
        left.download_button(
            "Download LIVE report", generated.live_bytes, generated.live_name,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        right.download_button(
            "Download client report", generated.client_bytes, generated.client_name,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.markdown("###### Email Text")
        st.code(_build_email_text(generated), language="text")


if __name__ == "__main__":
    run_app()
